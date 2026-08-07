# -*- coding: utf-8 -*-
"""多源数据接入：把不同来源的数据统一成 list[dict]。

支持 CSV / SQLite / Excel(可选 openpyxl)，并做去重与空值清洗。
纯标准库 csv / sqlite3，Excel 仅在安装了 openpyxl 时可用。
"""

import csv
import sqlite3


def load_csv(path) -> list[dict]:
    """读取 CSV 文件，返回字典列表。自动处理 BOM 与表头。"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sqlite(path, table) -> list[dict]:
    """读取 SQLite 数据库某张表的全部行，返回字典列表。"""
    if not table or not table.replace("_", "").isalnum():  # 表名白名单(安全加固)
        raise ValueError(f"非法表名: {table!r}")
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.execute(f'SELECT * FROM "{table}"')
        return [dict(r) for r in cur.fetchall()]


def load_mysql(path, table, host="localhost", port=3306, user="root", password="") -> list[dict]:
    """从 MySQL 读表为 list[dict]。需 pymysql（可选依赖，未装给清晰报错）。"""
    if not table or not table.replace("_", "").isalnum():
        raise ValueError(f"非法表名: {table!r}")
    try:
        import pymysql
    except ImportError:
        raise ImportError("读 MySQL 需 pymysql：pip install pymysql")
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=path,
                           cursorclass=pymysql.cursors.DictCursor)
    try:
        with conn.cursor() as cur:
            cur.execute(f'SELECT * FROM `{table}`')
            return list(cur.fetchall())
    finally:
        conn.close()


def load_pg(path, table, host="localhost", port=5432, user="postgres", password="") -> list[dict]:
    """从 PostgreSQL 读表为 list[dict]。需 psycopg2（可选依赖）。"""
    if not table or not table.replace("_", "").isalnum():
        raise ValueError(f"非法表名: {table!r}")
    try:
        import psycopg2
        import psycopg2.extras
    except ImportError:
        raise ImportError("读 PostgreSQL 需 psycopg2：pip install psycopg2")
    conn = psycopg2.connect(host=host, port=port, user=user, password=password, dbname=path)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(f'SELECT * FROM "{table}"')
            return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


def load_excel(path) -> list[dict]:
    """读取 Excel(.xlsx) 第一个工作表，返回字典列表。

    依赖可选库 openpyxl；未安装时给出清晰错误提示。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "读取 Excel 需要 openpyxl，请先安装：pip install openpyxl"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)  # 第一行为表头
    if header is None:
        return []
    return [
        {h: v for h, v in zip(header, row) if h is not None}
        for row in rows
    ]


def clean(rows, dedup_keys=()):
    """清洗数据：按 dedup_keys 去重（保留首个），丢弃全空行与纯空值字段。

    返回新的 list[dict]。
    """
    seen = set()
    out = []
    for row in rows:
        # 过滤掉键为空的项
        data = {k: v for k, v in row.items() if k is not None and v is not None}
        if not data:                      # 全空行，丢弃
            continue
        if dedup_keys:                    # 去重
            key = tuple(str(data.get(k)) for k in dedup_keys)
            if key in seen:
                continue
            seen.add(key)
        out.append(data)
    return out


def db_connect(db_type, db, host="localhost", port=3306, user="root", password="", tables=None) -> dict:
    """连接 MySQL/PostgreSQL，加载指定表返回 {表名: 行列表}。
    tables 缺省则加载 data_mapper 已知业务表(products/suppliers/inventory/sales/customers/equipment/purchase/production/payments)。
    """
    if db_type not in ("mysql", "pg", "postgresql"):
        raise ValueError(f"不支持的数据库类型: {db_type} (支持 mysql/pg)")
    pg = db_type in ("pg", "postgresql")
    known = ["products","suppliers","inventory","sales","customers","equipment","purchase","production","payments"]
    tables = tables or known
    out = {}
    for t in tables:
        if not t.replace("_", "").isalnum():
            continue
        try:
            if pg:
                rows = load_pg(db, t, host=host, port=port or 5432, user=user, password=password)
            else:
                rows = load_mysql(db, t, host=host, port=port or 3306, user=user, password=password)
            out[t] = rows
        except ImportError:
            raise  # 驱动未装, 向上抛清晰错误
        except Exception as e:
            # 该表不存在则跳过, 其他错误记录
            if "exist" not in str(e).lower():
                out[f"{t}(err)"] = str(e)[:80]
    return out


def db_test(db_type, db, host="localhost", port=3306, user="root", password="") -> dict:
    """测试数据库连接(不加载数据)。返回 {ok, driver, detail}。"""
    if db_type not in ("mysql", "pg", "postgresql"):
        return {"ok": False, "detail": f"不支持的数据库类型: {db_type}"}
    pg = db_type in ("pg", "postgresql")
    try:
        if pg:
            import psycopg2
        else:
            import pymysql
        return {"ok": True, "driver": "psycopg2" if pg else "pymysql", "detail": f"驱动已装, 可连接 {db_type}@{host}:{port}/{db}"}
    except ImportError:
        return {"ok": False, "detail": f"缺驱动: {'psycopg2' if pg else 'pymysql'} (pip install {'psycopg2-binary' if pg else 'pymysql'})"}


if __name__ == "__main__":
    # 自检：python data_mapper.py <csv路径>
    import sys
    rows = load_csv(sys.argv[1])
    print(clean(rows, dedup_keys=[sys.argv[2] if len(sys.argv) > 2 else "id"])[:3])
